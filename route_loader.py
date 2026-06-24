"""Le a planilha Excel da Rota 66 e popula a tabela 'waypoints' no SQLite.

Rode UMA vez (ou sempre que mudar a planilha):  python route_loader.py

Adaptado ao formato real da planilha (multi-aba):
  - Aba "Lista de Lugares": fonte principal dos pontos (Local, Cidade, Estado, Address).
    Como a planilha nao tem latitude/longitude, geocodifica pelo ENDERECO completo
    (bem mais preciso que pelo nome) via Nominatim.
  - Aba "POI": observacoes/dicas, casadas por cidade e anexadas a cada ponto.
"""
import asyncio
import datetime
import openpyxl
from config import EXCEL_PATH
from database import init_db, limpar_waypoints, inserir_waypoint, salvar_conhecimento
from services.geo import forward_geocode

ABA_LUGARES = "Lista de Lugares"
ABA_POI = "POI"

# Palavras que ajudam a classificar o tipo do ponto (so para o bot frasear melhor).
TIPOS = [
    ("aluguel de moto", ["eagle rider"]),
    ("hospedagem", ["inn", "hotel", "roof", "wyndham", "excalibur", "overland",
                    "days inn", "howard johnson", "motel", "lodge", "suites"]),
    ("aeroporto", ["airport", "(ord)", "aeroporto"]),
]


def classificar(nome):
    n = (nome or "").lower()
    for tipo, chaves in TIPOS:
        if any(c in n for c in chaves):
            return tipo
    return "ponto de interesse"


def _carregar_poi(wb):
    """Monta um dicionario {cidade_minuscula: 'dica1; dica2; ...'} a partir da aba POI."""
    if ABA_POI not in wb.sheetnames:
        return {}
    ws = wb[ABA_POI]
    dicas = {}
    for linha in list(ws.iter_rows(values_only=True))[1:]:  # pula cabecalho
        # Colunas: Dia, Cidade, Local, Observacao
        cidade = (linha[1] or "").strip() if len(linha) > 1 and linha[1] else ""
        local = (str(linha[2]).strip() if len(linha) > 2 and linha[2] else "")
        obs = (str(linha[3]).strip() if len(linha) > 3 and linha[3] else "")
        if not cidade:
            continue
        texto = " - ".join(p for p in (local, obs) if p)
        if texto:
            dicas.setdefault(cidade.lower(), []).append(texto)
    return {c: "; ".join(v) for c, v in dicas.items()}


def _fmt(v):
    """Formata valores para texto legivel (datas e duracoes viram texto)."""
    if isinstance(v, datetime.datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, datetime.timedelta):
        h = v.seconds // 3600
        m = (v.seconds % 3600) // 60
        return f"{h}h{m:02d}"
    return "" if v is None else str(v).strip()


def _dump_planilha(wb):
    """Le TODAS as abas e devolve um texto unico com o conteudo, para o Gemini
    poder responder perguntas sobre qualquer tema da planilha (datas, custos,
    hoteis, distancias, etc.)."""
    blocos = []
    for nome in wb.sheetnames:
        ws = wb[nome]
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            continue
        bloco = [f"## Aba: {nome}"]
        ultima_data = None
        for linha in linhas:
            cels = list(linha)
            # Na aba "Roteiro" a data so aparece na 1a perna do dia; replicamos
            # para baixo para cada trecho carregar a sua data explicitamente.
            if nome == "Roteiro" and len(cels) > 1:
                if isinstance(cels[1], datetime.datetime):
                    ultima_data = cels[1]
                elif cels[1] is None and ultima_data and any(c is not None for c in cels[2:]):
                    cels[1] = ultima_data
            valores = [_fmt(c) for c in cels]
            if not any(valores):  # linha totalmente vazia
                continue
            bloco.append(" | ".join(valores))
        blocos.append("\n".join(bloco))
    return "\n\n".join(blocos)


async def carregar():
    init_db()
    limpar_waypoints()

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if ABA_LUGARES not in wb.sheetnames:
        print(f"Nao achei a aba '{ABA_LUGARES}'. Abas disponiveis: {wb.sheetnames}")
        return

    poi_por_cidade = _carregar_poi(wb)
    ws = wb[ABA_LUGARES]
    linhas = list(ws.iter_rows(values_only=True))
    # Cabecalho esperado: #, Local, Cidade, Estado, Address, POI Google
    print("Cabecalho:", linhas[0])

    total = 0
    for i, linha in enumerate(linhas[1:], start=1):
        # Desempacota com segurranca (linhas podem ter menos colunas)
        col = list(linha) + [None] * (6 - len(linha))
        ordem, nome, cidade, estado, endereco, link = col[0], col[1], col[2], col[3], col[4], col[5]

        if not nome:  # linhas em branco (ex.: # 3 e 4) sao ignoradas
            continue

        # Geocodifica pelo endereco completo; se faltar, usa nome + cidade + estado.
        alvo = endereco or ", ".join(p for p in (str(nome), cidade, estado, "USA") if p)
        print(f"  Geocodificando: {str(nome)[:40]} ...")
        lat, lon = await forward_geocode(alvo)
        if lat is None:
            print(f"    [aviso] nao geocodificou '{nome}'. Vai ficar sem coordenada.")

        try:
            ordem_int = int(ordem)
        except (ValueError, TypeError):
            ordem_int = i

        dica = poi_por_cidade.get((cidade or "").lower(), "")

        inserir_waypoint(
            ordem=ordem_int,
            nome=str(nome),
            tipo=classificar(nome),
            cidade=str(cidade or ""),
            lat=lat,
            lon=lon,
            dicas=dica,
            link=str(link or ""),
        )
        total += 1

    # Guarda o texto de TODAS as abas para as perguntas em linguagem natural.
    salvar_conhecimento(_dump_planilha(wb))

    print(f"\nPronto! {total} pontos carregados. {len(poi_por_cidade)} cidades com dicas da aba POI.")
    print("Conteudo de todas as abas indexado para perguntas livres.")


if __name__ == "__main__":
    asyncio.run(carregar())
